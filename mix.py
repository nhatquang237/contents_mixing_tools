# -*- coding: utf-8 -*-
"""
Created on Tue Jun  1 15:32:18 2021

@author: NhatQuang
"""

import random
import os
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
   
def tem_list(A, B):
    tem_list = list()
    for i in A:
        if(i!=B):
            tem_list.append(i)
    return tem_list

def random_pick(collection,slot):
    result = list()
    tempory_list = collection
    count = 0
    while count<slot:
        result.append(random.choice(tempory_list))
        tempory_list = tem_list(tempory_list, result[count])
        count+=1
    return result

def file_open():
    global link
    root.filename = filedialog.askopenfilename(initialdir = "/",
                                               title = "Select File",
                                               filetype = (('xlsx', '*.xlsx'),
                                                           ('xls', '*.xls')))
    
    link_box.config(state='normal')
    link_box.delete(0, END)    
    link_box.insert(0, root.filename)
    
    link = root.filename.replace('/', '\\\\')   
    link_box.config(state='disabled')


def process():
    #Gather information from wookbook
    wb = load_workbook(link)
    ws = wb.active    
    
    slot_title = ws.cell(1,4).value
    slot_content = ws.cell(2,4).value
    size_titles = ws.cell(3,4).value
    size_contents = ws.cell(4,4).value
    size_combination = ws.cell(5,4).value
    
    
    if size_combination > size_titles*size_contents:
        size_combination = size_titles*size_contents
    
    titles = list()
    for i in range(2,size_titles+2):
        titles.append(ws.cell(i,1).value)
    contents = list()
    for i in range(2,size_contents+2):
        contents.append(ws.cell(i,2).value)

    unique_list = list()
    while( len(unique_list)<size_combination):
        result_C = random_pick(titles,slot_title)
        result_N = random_pick(contents,slot_content)
        if(result_C+result_N) in unique_list:
            pass
        else:
            unique_list.append(result_C+result_N)
    
    for i in range(size_combination):
        for k in range(slot_content+slot_title):
            ws.cell(i+1,k+6).value = unique_list[i][k]
    wb.save(link)
    os.startfile(link)
        

###### GUI design

root = Tk()
root.resizable(width = 0, height = 0)
root.title("Copyright 2021 © Eng.Nhat Quang")
root.minsize(450, 70)

 
link_box = Entry(root, width=60, relief=FLAT, highlightbackground="black", highlightthickness=1)
link_box.grid(row = 0, column = 1, padx = 5, ipady = 2, sticky = "W", columnspan = 6)

button_1 = Button(root, text = 'Open File',
command=file_open,
background = 'SteelBlue1', borderwidth=4).grid(row = 0, 
column = 0, padx = 10, pady = 2, ipady = 2,
sticky = "WE")

button_2 = Button(root, command=process, text = ' RUN ',
background = 'IndianRed3', highlightbackground="black",
highlightthickness=1, borderwidth=1).grid(row = 20,
column = 0, ipadx = 5, columnspan = 8)
                                               
root.mainloop() 